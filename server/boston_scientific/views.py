from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser

import pandas as pd
from datetime import datetime
from boston_scientific.models import Producto, Cliente, ArchivoDeCarga
import uuid
from .serializers import (
    ProductoSerializer,
    ClienteSerializer,
    ArchivoDeCargaSerializer,
    ProductoCargaSerializer,
)
import openpyxl
from django.http import HttpResponse
from io import BytesIO


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer


class ArchivoDeCargaViewSet(viewsets.ModelViewSet):
    queryset = ArchivoDeCarga.objects.all()
    serializer_class = ArchivoDeCargaSerializer


class CargaMasivaProductosView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        archivo = request.FILES.get("file")
        if not archivo or not archivo.name.endswith(".xlsx"):
            return Response({"error": "Se requiere un archivo .xlsx válido."},
                            status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active

            # Validar encabezados
            encabezados = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            if encabezados[:3] != ["GTIN", "Código", "Descripción"]:
                return Response(
                    {"error": "Los encabezados deben ser: GTIN, Código, Descripción"},
                    status=status.HTTP_400_BAD_REQUEST
                )

            creados = 0
            actualizados = 0
            productos_cargados = []

            for fila in ws.iter_rows(min_row=2, values_only=True):
                gtin, codigo, descripcion = fila[:3]
                if not gtin or not codigo:
                    continue

                obj, created = Producto.objects.update_or_create(
                    gtin=str(gtin).strip(),
                    defaults={
                        "codigo": str(codigo).strip(),
                        "descripcion": str(descripcion).strip() if descripcion else "",
                    }
                )

                productos_cargados.append({
                    "gtin": obj.gtin,
                    "codigo": obj.codigo,
                    "descripcion": obj.descripcion,
                    "accion": "creado" if created else "actualizado"
                })

                if created:
                    creados += 1
                else:
                    actualizados += 1

            return Response({
                "mensaje": f"Productos cargados correctamente.",
                "creados": creados,
                "actualizados": actualizados,
                "productos": productos_cargados
            })

        except Exception as e:
            return Response({"error": f"Error procesando el archivo: {str(e)}"},
                            status=status.HTTP_400_BAD_REQUEST)


# views.py (fragmento modificado de CargarTXTView)

class CargarTXTView(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request):
        archivo = request.FILES.get("archivo")
        cliente_codigo = request.data.get("cliente_codigo")

        if not archivo:
            return Response({"error": "No se proporcionó un archivo."}, status=400)

        if not cliente_codigo:
            return Response({"error": "Se requiere el código del cliente."}, status=400)

        cliente = Cliente.objects.filter(codigo=cliente_codigo).first()
        if not cliente:
            return Response({"error": "Cliente no encontrado."}, status=404)

        try:
            lines = archivo.read().decode("utf-8").splitlines()
        except Exception:
            return Response({"error": "No se pudo leer el archivo."}, status=400)

        delivery_number = None
        date = None
        productos = []
        gtin_no_encontrados = set()

        for line in lines:
            if line.startswith("DeliveryNumber"):
                delivery_number = line.split(":")[1].strip()
            elif line.startswith("Date"):
                date = line.split(":")[1].strip()

        product_lines = [l for l in lines if l.startswith("Product")]

        for line in product_lines:
            try:
                product_data = line.split(",")[0].split(":")[1]
                tag_data = line.split(",")[1].split(":")[1]
                gtin = product_data[2:16]
                usebydate = product_data[18:24]
                batch = product_data[26:]

                caducidad = f"20{usebydate[:2]}-{usebydate[2:4]}-{usebydate[4:]}"
                fecha_reposicion = date

                producto_obj = Producto.objects.filter(gtin=gtin).first()
                if not producto_obj:
                    gtin_no_encontrados.add(gtin)
                    continue

                productos.append({
                    "GTIN": gtin,
                    "Codigo": producto_obj.codigo,
                    "Lote": batch,
                    "Caducidad": caducidad,
                    "Documento de reposicion": delivery_number,
                    "Fecha de reposicion": fecha_reposicion,
                    "No. de envio": delivery_number,
                    "Orden de compra": "S/O",
                    "Ticket de salida": "S/T",
                    "Almacen BSCI": "CEDIS",
                    "No. de cliente": cliente_codigo,  
                    "Etiqueta RFID": tag_data,
                })

            except Exception as e:
                print(f"Error procesando línea: {line} | {str(e)}")

        numero_archivo = str(uuid.uuid4())

        if gtin_no_encontrados:
            return Response({
                "errores_gtin": list(gtin_no_encontrados),
                "numero_archivo": numero_archivo,
                "datos": productos
            }, status=status.HTTP_207_MULTI_STATUS)

        return Response({"datos": productos, "numero_archivo": numero_archivo})




class GenerarExcelView(APIView):
    def post(self, request, numero_archivo):
        datos = request.data.get("datos", [])
        if not datos:
            return Response({"error": "No se enviaron datos."}, status=400)

        # Guardar en base de datos
        for producto in datos:
            ArchivoDeCarga.objects.create(
                numero_archivo=numero_archivo,
                codigo=producto["Codigo"],
                lote=producto["Lote"],
                caducidad=producto["Caducidad"],
                documento_reposicion=producto["Documento de reposicion"],
                fecha_reposicion=producto["Fecha de reposicion"],
                numero_envio=producto["No. de envio"],
                orden_compra=producto["Orden de compra"],
                ticket_salida=producto["Ticket de salida"],
                almacen_bsci=producto["Almacen BSCI"],
                numero_cliente=producto["No. de cliente"],
                etiqueta_rfid=producto["Etiqueta RFID"],
            )

        # Luego genera el Excel
        registros = ArchivoDeCarga.objects.filter(numero_archivo=numero_archivo)
        df = pd.DataFrame(list(registros.values(
            "codigo", "lote", "caducidad", "documento_reposicion", "fecha_reposicion",
            "numero_envio", "orden_compra", "ticket_salida", "almacen_bsci",
            "numero_cliente", "etiqueta_rfid"
        )))
        df.columns = [
            "Código", "Lote", "Caducidad", "Documento de reposicion", "Fecha de reposicion",
            "No. de envio", "Orden de compra", "Ticket de salida", "Almacen BSCI",
            "No. de cliente", "Etiqueta RFID"
        ]

        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Carga")
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=archivo_{numero_archivo}.xlsx'
        return response
